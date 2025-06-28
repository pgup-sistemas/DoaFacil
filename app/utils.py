import csv
import io
from datetime import datetime
from weasyprint import HTML
from flask import render_template, current_app
from app.models import Lista, Item, Doacao
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_lista_csv(lista):
    """Exporta uma lista para CSV (completo para administradores)"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Cabeçalho
    writer.writerow(['Relatório Completo de Doações'])
    writer.writerow([f'Nome da Lista: {lista.nome}'])
    writer.writerow([f'Unidade: {lista.unidade.nome}'])
    writer.writerow([f'Data de Criação: {lista.data_criacao.strftime("%d/%m/%Y %H:%M") }'])
    writer.writerow([''])
    writer.writerow(['Tipo', 'Doador', 'CPF', 'Contato', 'CEP', 'Item', 'Quantidade/Valor', 'Data', 'Observação'])
    
    for item in lista.itens:
        for doacao in item.doacoes:
            if doacao.tipo_doacao == 'item':
                # Doação física
                quantidade_valor = f"{doacao.quantidade} {item.unidade_medida}"
            else:
                # Doação financeira
                quantidade_valor = f"R$ {doacao.valor_dinheiro:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            writer.writerow([
                'Física' if doacao.tipo_doacao == 'item' else 'Financeira',
                doacao.doador_nome or '',
                doacao.cpf or '',
                doacao.contato or '',
                doacao.cep or '',
                item.nome,
                quantidade_valor,
                doacao.data_doacao.strftime('%d/%m/%Y %H:%M') if doacao.data_doacao else '',
                doacao.observacao or ''
            ])
    return output.getvalue()

def export_lista_pdf(lista):
    """Exporta uma lista para PDF"""
    # Renderizar template HTML
    html_content = render_template('export/pdf_lista.html', lista=lista, datetime=datetime)
    
    # Gerar PDF
    pdf = HTML(string=html_content).write_pdf()
    return pdf

def format_currency(value):
    """Formata valor como moeda brasileira"""
    return f"R$ {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def get_progress_color(percentual):
    """Retorna a cor baseada no percentual de conclusão"""
    if percentual >= 100:
        return 'is-success'
    elif percentual >= 75:
        return 'is-warning'
    elif percentual >= 50:
        return 'is-info'
    else:
        return 'is-danger'

def send_notification_email(lista, doacao):
    """Envia e-mail de notificação para o administrador"""
    try:
        from flask_mail import Message
        from app import mail
        
        if not lista.unidade.email:
            return False
            
        msg = Message(
            f'Nova doação na lista "{lista.nome}"',
            sender=current_app.config['MAIL_USERNAME'],
            recipients=[lista.unidade.email]
        )
        
        msg.body = f"""
        Nova doação registrada!
        
        Lista: {lista.nome}
        Doador: {doacao.doador_nome}
        Item: {doacao.item.nome}
        Quantidade: {doacao.quantidade} {doacao.item.unidade_medida}
        Data: {doacao.data_doacao.strftime("%d/%m/%Y %H:%M")}
        
        Acesse o dashboard para mais detalhes.
        """
        
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Erro ao enviar e-mail: {e}")
        return False

def generate_whatsapp_share_url(lista):
    """Gera URL para compartilhamento via WhatsApp"""
    url = f"https://wa.me/?text=Ajude%20nossa%20arrecadação!%20{lista.nome}%20-%20"
    return url

def generate_whatsapp_share_url_with_unidade(lista):
    """Gera URL para compartilhamento via WhatsApp com informações da unidade"""
    try:
        base_text = f"Ajude nossa arrecadação! {lista.nome}"
        
        # Adicionar informações da unidade se disponível
        if lista.unidade:
            # Verificar se a unidade tem o atributo whatsapp_numero (pode não existir em unidades antigas)
            if hasattr(lista.unidade, 'whatsapp_numero') and lista.unidade.whatsapp_numero:
                # Se a unidade tem WhatsApp configurado, usar o número dela
                return f"https://wa.me/{lista.unidade.whatsapp_numero}?text={base_text}%20-%20"
            elif lista.unidade.nome:
                # Se não tem WhatsApp, incluir o nome da unidade na mensagem
                base_text += f" - {lista.unidade.nome}"
        
        # URL padrão
        return f"https://wa.me/?text={base_text}%20-%20"
    except Exception as e:
        # Fallback para URL padrão em caso de erro
        return f"https://wa.me/?text=Ajude%20nossa%20arrecadação!%20{lista.nome}%20-%20"

def generate_facebook_share_url(lista):
    """Gera URL para compartilhamento via Facebook"""
    return f"https://www.facebook.com/sharer/sharer.php?u="

def generate_facebook_share_url_with_unidade(lista):
    """Gera URL para compartilhamento via Facebook com informações da unidade"""
    try:
        base_url = f"https://www.facebook.com/sharer/sharer.php?u="
        
        # Se a unidade tem página do Facebook configurada, adicionar hashtag
        if lista.unidade and hasattr(lista.unidade, 'facebook_url') and lista.unidade.facebook_url:
            # Extrair nome da página do Facebook da URL
            import re
            match = re.search(r'facebook\.com/([^/?]+)', lista.unidade.facebook_url)
            if match:
                page_name = match.group(1)
                return f"{base_url}&hashtag=%23{page_name}"
        
        return base_url
    except Exception as e:
        # Fallback para URL padrão em caso de erro
        return f"https://www.facebook.com/sharer/sharer.php?u="

def get_unidade_social_links(lista):
    """Retorna links das redes sociais da unidade para exibição"""
    try:
        if not lista.unidade:
            return {}
        
        links = {}
        
        # Verificar se a unidade tem os novos campos antes de tentar acessá-los
        if hasattr(lista.unidade, 'whatsapp_numero') and lista.unidade.whatsapp_numero:
            links['whatsapp'] = f"https://wa.me/{lista.unidade.whatsapp_numero}"
        
        if hasattr(lista.unidade, 'facebook_url') and lista.unidade.facebook_url:
            links['facebook'] = lista.unidade.facebook_url
        
        if hasattr(lista.unidade, 'instagram_url') and lista.unidade.instagram_url:
            links['instagram'] = lista.unidade.instagram_url
        
        if hasattr(lista.unidade, 'website_url') and lista.unidade.website_url:
            links['website'] = lista.unidade.website_url
        
        return links
    except Exception as e:
        # Retornar dicionário vazio em caso de erro
        return {}

def export_lista_excel(lista):
    wb = Workbook()
    
    # Aba 1: Resumo dos Itens
    ws_resumo = wb.active
    ws_resumo.title = "Resumo dos Itens"

    # Cabeçalho
    ws_resumo.append([f"Lista: {lista.nome}"])
    ws_resumo.append([f"Unidade: {lista.unidade.nome}"])
    ws_resumo.append([f"Descrição: {lista.descricao or ''}"])
    ws_resumo.append([f"Data de Criação: {lista.data_criacao.strftime('%d/%m/%Y')}"])
    ws_resumo.append([f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws_resumo.append([])

    # Títulos das colunas
    ws_resumo.append(["Item", "Descrição", "Necessário", "Arrecadado", "Restante", "Unidade", "Progresso (%)", "Valor PIX"])

    # Dados dos itens
    for item in lista.itens:
        ws_resumo.append([
            item.nome,
            item.descricao or '',
            item.quantidade_necessaria,
            item.quantidade_arrecadada,
            item.quantidade_restante,
            item.unidade_medida,
            round(item.percentual_conclusao, 1),
            f"R$ {item.valor_financeiro_arrecadado:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if item.valor_financeiro_arrecadado > 0 else '-'
        ])

    # Ajustar largura das colunas
    for col in ws_resumo.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_resumo.column_dimensions[column].width = max(12, max_length + 2)

    # Aba 2: Detalhamento das Doações
    ws_doacoes = wb.create_sheet("Doações Detalhadas")
    
    # Cabeçalho das doações
    ws_doacoes.append([f"Lista: {lista.nome} - Detalhamento de Doações"])
    ws_doacoes.append([f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws_doacoes.append([])
    ws_doacoes.append(["Tipo", "Doador", "CPF", "Contato", "CEP", "Item", "Quantidade/Valor", "Data", "Observação"])
    
    # Dados das doações
    for item in lista.itens:
        for doacao in item.doacoes:
            if doacao.tipo_doacao == 'item':
                # Doação física
                quantidade_valor = f"{doacao.quantidade} {item.unidade_medida}"
            else:
                # Doação financeira
                quantidade_valor = f"R$ {doacao.valor_dinheiro:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            ws_doacoes.append([
                'Física' if doacao.tipo_doacao == 'item' else 'Financeira',
                doacao.doador_nome or '',
                doacao.cpf or '',
                doacao.contato or '',
                doacao.cep or '',
                item.nome,
                quantidade_valor,
                doacao.data_doacao.strftime('%d/%m/%Y %H:%M') if doacao.data_doacao else '',
                doacao.observacao or ''
            ])
    
    # Ajustar largura das colunas da aba de doações
    for col in ws_doacoes.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_doacoes.column_dimensions[column].width = max(12, max_length + 2)

    # Salvar em bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def consultar_cep(cep):
    """
    Consulta CEP via API ViaCEP
    Retorna dados do endereço ou None se erro
    """
    try:
        # Limpar CEP (remover caracteres especiais)
        cep_limpo = re.sub(r'[^\d]', '', cep)
        
        if len(cep_limpo) != 8:
            return None
        
        # Consultar ViaCEP
        url = f'https://viacep.com.br/ws/{cep_limpo}/json/'
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            
            # Verificar se não há erro
            if not data.get('erro'):
                return {
                    'cep': data.get('cep', ''),
                    'logradouro': data.get('logradouro', ''),
                    'bairro': data.get('bairro', ''),
                    'cidade': data.get('localidade', ''),
                    'estado': data.get('uf', ''),
                    'ibge': data.get('ibge', ''),
                    'ddd': data.get('ddd', ''),
                    'success': True
                }
        
        return None
        
    except requests.RequestException as e:
        print(f"Erro na requisição CEP: {e}")
        return None
    except Exception as e:
        print(f"Erro ao consultar CEP: {e}")
        return None

def formatar_cep(cep):
    """
    Formata CEP no padrão 00000-000
    """
    cep_limpo = re.sub(r'[^\d]', '', cep)
    if len(cep_limpo) == 8:
        return f"{cep_limpo[:5]}-{cep_limpo[5:]}"
    return cep

def validar_cep(cep):
    """
    Valida formato do CEP
    """
    cep_limpo = re.sub(r'[^\d]', '', cep)
    return len(cep_limpo) == 8 and cep_limpo.isdigit()

def export_unidade_completa(unidade):
    """
    Exporta dados completos de uma unidade organizadora em Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados da Unidade"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    section_font = Font(bold=True, color="667eea", size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título principal
    ws.merge_cells('A1:H1')
    ws['A1'] = f"RELATÓRIO COMPLETO - {unidade.nome.upper()}"
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Informações Básicas
    ws['A3'] = "INFORMAÇÕES BÁSICAS"
    ws['A3'].font = section_font
    
    basic_info = [
        ["Nome da Organização", unidade.nome],
        ["Nome Fantasia", unidade.nome_fantasia or "Não informado"],
        ["Razão Social", unidade.razao_social or "Não informado"],
        ["Responsável", unidade.responsavel],
        ["Email Principal", unidade.email or "Não informado"],
        ["Email Secundário", unidade.email_secundario or "Não informado"],
        ["Descrição", unidade.descricao or "Não informado"]
    ]
    
    for i, (label, value) in enumerate(basic_info, 4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = border
        ws[f'B{i}'].border = border
    
    # Dados Legais
    ws['A12'] = "DADOS LEGAIS/INSTITUCIONAIS"
    ws['A12'].font = section_font
    
    legal_info = [
        ["CNPJ", unidade.cnpj_formatado or "Não informado"],
        ["Tipo de Organização", unidade.tipo_organizacao.title() if unidade.tipo_organizacao else "Não informado"],
        ["Data de Fundação", unidade.data_fundacao.strftime("%d/%m/%Y") if unidade.data_fundacao else "Não informado"],
        ["Registro/Inscrição", unidade.registro_inscricao or "Não informado"]
    ]
    
    for i, (label, value) in enumerate(legal_info, 13):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = border
        ws[f'B{i}'].border = border
    
    # Endereço
    ws['A18'] = "ENDEREÇO COMPLETO"
    ws['A18'].font = section_font
    
    address_info = [
        ["CEP", unidade.cep or "Não informado"],
        ["Logradouro", unidade.logradouro or "Não informado"],
        ["Número", unidade.numero or "Não informado"],
        ["Complemento", unidade.complemento or "Não informado"],
        ["Bairro", unidade.bairro or "Não informado"],
        ["Cidade", unidade.cidade or "Não informado"],
        ["Estado", unidade.estado or "Não informado"],
        ["Endereço Completo", unidade.endereco_completo]
    ]
    
    for i, (label, value) in enumerate(address_info, 19):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = border
        ws[f'B{i}'].border = border
    
    # Contatos
    ws['A28'] = "CONTATOS"
    ws['A28'].font = section_font
    
    contact_info = [
        ["Telefone Fixo", unidade.telefone_fixo or "Não informado"],
        ["WhatsApp", unidade.whatsapp_numero or "Não informado"],
        ["Horário de Funcionamento", unidade.horario_funcionamento or "Não informado"],
        ["Dias de Funcionamento", unidade.dias_funcionamento or "Não informado"]
    ]
    
    for i, (label, value) in enumerate(contact_info, 29):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = border
        ws[f'B{i}'].border = border
    
    # Presença Digital
    ws['A34'] = "PRESENÇA DIGITAL"
    ws['A34'].font = section_font
    
    digital_info = [
        ["Website", unidade.website_url or "Não informado"],
        ["Facebook", unidade.facebook_url or "Não informado"],
        ["Instagram", unidade.instagram_url or "Não informado"],
        ["LinkedIn", unidade.linkedin_url or "Não informado"],
        ["YouTube", unidade.youtube_url or "Não informado"]
    ]
    
    for i, (label, value) in enumerate(digital_info, 35):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = border
        ws[f'B{i}'].border = border
    
    # Dados para Relatórios
    ws['A41'] = "DADOS PARA RELATÓRIOS"
    ws['A41'].font = section_font
    
    report_info = [
        ["Categoria", unidade.categoria.title() if unidade.categoria else "Não informado"],
        ["Público-Alvo", unidade.publico_alvo or "Não informado"],
        ["Área de Atuação", unidade.area_atuacao.title() if unidade.area_atuacao else "Não informado"],
        ["Tamanho da Organização", unidade.tamanho_organizacao.title() if unidade.tamanho_organizacao else "Não informado"]
    ]
    
    for i, (label, value) in enumerate(report_info, 42):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = border
        ws[f'B{i}'].border = border
    
    # Listas da Unidade
    ws['A47'] = "LISTAS DA UNIDADE"
    ws['A47'].font = section_font
    
    # Cabeçalho da tabela de listas
    listas_headers = ["Nome", "Modo", "Status", "Itens", "Arrecadado", "Progresso", "Data Criação"]
    for i, header in enumerate(listas_headers, 48):
        cell = ws[f'{get_column_letter(i+1)}{48}']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Dados das listas
    for i, lista in enumerate(unidade.listas, 49):
        ws[f'A{i}'] = lista.nome
        ws[f'B{i}'] = lista.modo.title()
        ws[f'C{i}'] = "Ativa" if lista.ativa and not lista.concluida and not lista.cancelada else "Concluída" if lista.concluida else "Cancelada"
        ws[f'D{i}'] = len(lista.itens)
        ws[f'E{i}'] = lista.total_arrecadado
        ws[f'F{i}'] = f"{lista.percentual_conclusao:.1f}%"
        ws[f'G{i}'] = lista.data_criacao.strftime("%d/%m/%Y")
        
        for col in range(1, 8):
            ws[f'{get_column_letter(col)}{i}'].border = border
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue() 